import openpyxl

def getPath(week):
    if week == '_верхній_':
        return r'Database\Schedule (Up).xlsx'
    elif week == '_нижній_':
        return r'Database\Schedule (Down).xlsx'

def getColumns(day):
    columnName = day*2
    return columnName


def getScheduleForDay(week, day):
    path = getPath(week)
    wbObj = openpyxl.load_workbook(path)
    sheetObj = wbObj.active
    schedule = ''
    columns = [1]
    columns.append(getColumns(day))
    columns.append(getColumns(day)+1)
    for i in range(1, 7):
        for j in columns:
            cellObj = sheetObj.cell(row=i,column=j)
            val = cellObj.value
            if val==None or val=='-' or (i==1 and j==1):
                val=' '
            schedule+=str(val)+' '
        schedule+='\n'
    return schedule

def getScheduleForWeek(week):
    schedule = ''
    for i in range(1, 8):
        schedule+=getScheduleForDay(week, i)
        schedule+='----------------------------------------\n\n'
    return schedule

def getScheduleOfCalls():
    path = r'Schedules/Schedule (Calls).xlsx'
    wbObj = openpyxl.load_workbook(path)
    sheetObj = wbObj.active
    schedule = ''
    for i in range(1,10):
        k = 1
        for j in range(1,5):
            if j==3:
                schedule+='- '
            else:
                cellObj = sheetObj.cell(row=i, column=k)
                val = str(cellObj.value)[0:5]
                schedule += str(val) + ' '
                k+=1
        schedule += '\n'
    return schedule